
from openpyxl.styles.fonts import Font
from time import sleep
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

##################################################################
# Excel Input & Ouput
##################################################################

class ExcelIo:

    def __init__(self) -> None:
        pass

    ##################################################################
    # Excel File Write
    ##################################################################
 
    def write(self, file_name, data):    
        f = openpyxl.Workbook()
        # sheet = f.worksheets[0]
        
        sheet = f.active
        sheet.title = 'Data'
        # sheet['A1:C1'].font = Font(size=48, italic=True, color='ff9999')

        data_length = len(data)

        for index, value in enumerate(data):
            sheet.append(value)
            print('\rExcel 작성 중 : {}'.format(round((index+1)/data_length*100)), end='')
            # sleep(0.002)
            
        print('\n')

        f.save(f"./5_excel_db(xls)/{file_name}.xlsx")

    ##################################################################
    # Excel File Read
    ##################################################################

    def read(self, file_name):
        fr = load_workbook(f"./5_excel_db(xls)/{file_name}.xlsx", read_only=True, data_only=True)

        data = list()

        range = fr['Data']
        max_row = range.max_row

        for index, row in enumerate(range):
            row_data = list()
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
            print('\rExcel 읽는 중 : {}'.format(round((index+1)/max_row*100)), end='')
            # sleep(0.002)

        print('\n')
        
        return data

    ##################################################################
    # Excel File Read
    ##################################################################

    def print(self, file_name):
        fr = load_workbook(f"./5_excel_db(xls)/{file_name}.xlsx", read_only=True, data_only=True)

        range = fr['Data']
        for row in range:
            for cell in row:
                print(cell.value, end= ' | ')
            print('\n')

        # 이게 더 느린데..?
        # sheet = fr['Data']
        
        # max_row = sheet.max_row
        # max_column = sheet.max_column

        # for i in range(1, max_row + 1):
        #     for j in range(1, max_column + 1):
        #         print(sheet.cell(row=i, column=j).value, end= ' | ')
        #     print('\n')


        