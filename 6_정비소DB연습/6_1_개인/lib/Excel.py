
from openpyxl.styles.fonts import Font
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.alignment import Alignment

from time import sleep
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

##################################################################
# Excel Input & Ouput
##################################################################

class ExcelIo:

    def __init__(self) -> None:
        pass

    ##################################################################
    # Excel File - Title Style
    ##################################################################

    def title_style(self, sheet):
        for num in range(len(sheet["1"])):
            sheet.cell(1, num+1).font = Font(size=11, color='f2f2f2')
            sheet.cell(1,num+1).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            sheet.cell(1,num+1).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(1,num+1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ##################################################################
    # Excel File - All Columns Auto Fit
    ##################################################################

    def auto_fit_columns(self, sheet):
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length * 1.75 * 1.2

    ##################################################################
    # Excel File Write
    ##################################################################
 
    def write(self, file_name, data):    
        f = openpyxl.Workbook()
        # sheet = f.worksheets[0]
        
        sheet = f.active
        sheet.title = 'Data'

        data_length = len(data)

        for index, value in enumerate(data):
            sheet.append(value)
            print('\rExcel 작성 중 : {}%'.format(round((index+1)/data_length*100)), end='')
            # sleep(0.002)
        
        ExcelIo.title_style(self, sheet)
        ExcelIo.auto_fit_columns(self, sheet)
            
        print('\n')

        f.save(file_name)

    ##################################################################
    # Excel File Read
    ##################################################################

    def read(self, file_name):
        load_wb = load_workbook(file_name, read_only=True, data_only=True)

        data = list()

        load_ws = load_wb['전시장']
        max_row = load_ws.max_row

        for index, row in enumerate(load_ws):
            row_data = list()
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
            print('\rExcel 읽는 중 : {}%'.format(round((index+1)/max_row*100)), end='')

        print('\n')
        
        return data

    ##################################################################
    # Excel File Read
    ##################################################################

    def print(self, file_name):
        fr = load_workbook(file_name, read_only=True, data_only=True)

        range = fr['Data']
        for row in range:
            for cell in row:
                print(cell.value, end= ' | ')
            print('\n')

        # 이게 더 느린데..?
        # sheet = fr['Data']
        
        # max_row = sheet.max_row
        # max_column = sheet.max_column

        # for i in range(1, max_row + 1):
        #     for j in range(1, max_column + 1):
        #         print(sheet.cell(row=i, column=j).value, end= ' | ')
        #     print('\n')
